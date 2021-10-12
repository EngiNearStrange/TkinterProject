from tkinter import *
import tkinter.messagebox as tmsg
# import pandas as pd
import openpyxl
# from openpyxl import workbook
# import pathlib

#Logic and geometry definition
root = Tk()
root.geometry("980x580")
root.minsize(800, 550)
root.title("WFM Assistance Tool - by Ranjeet Varma (@EngiNearStrange)")

#Defining button functions
def rosterfunc():
    print("This will switch to Roster request window")

def masterfunc():
    print("This will switch to Master data edit window")

fbkvalue = ''
new_window = ''
def feedbackfunc():
    global fbkvalue
    global new_window
    new_window = Toplevel(root)
    new_window.geometry("400x250")
    new_window.title("Feedback")
    fbk = Label(new_window, text = "Please type your feedback in the box below", padx = 10, pady = 20)
    fbk.grid(row = 0, column = 0)
    fbkvalue = StringVar()
    fbkentry = Entry(new_window, textvariable = fbkvalue, width = 40)
    fbkentry.grid(row = 1, column = 0, padx = 10, pady = 10)
    Button(new_window, text = "Submit Feedback", bg = "blue", fg = "white", command = submitfeedback).grid(row = 2, column = 0, padx = 1, pady = 15)
    Button(new_window, text = "Clear", bg = "blue", fg = "white", padx = 4, command = lambda: fbkentry.delete("0", "end")).grid(row = 1, column = 1, padx = 5, pady = 15)
    Button(new_window, text = "Cancel", bg = "blue", fg = "white", command = lambda: new_window.destroy()).grid(row = 2, column = 1, padx = 5, pady = 15)

def submitfeedback():
    print("Submitting Feedback")
    with open("Feedback.txt", "a") as fbk:
        fbk.write("\n" + fbkvalue.get())
    print(f"The feedback was {fbkvalue.get()} and has been submitted")
    tmsg.showinfo("Feedback submitted", "Your feedback has been recorded, Thank You!!")
    new_window.destroy()

def aboutfunc():
    tmsg.showinfo("About","Developed for Accenture Services Pvt. Ltd. by \nRanjeet Varma - @EngiNearStrange")

def retdatafunc():
    print("This will retrieve existing alignment data from server")

def updatafunc():
    # data1 = [{"Sap ID": empidvalue.get(), "SME": smevalue.get(), "SME EID": smeeidvalue.get(),
    #           "Supervisor": supvalue.get(), "Sup Enterprise ID": supeidvalue.get(), "Shift Lead": slvalue.get(),
    #           "Shift Lead EID": sleidvalue.get(), "Ops Lead": olvalue.get(), "Ops Lead EID": oleidvalue.get()}]
    # df1 = pd.DataFrame(data1)
    # df1.to_excel('new1.csv', "a", index = False)
    # data1 = [{}]
    # print(df1)
    a = empidvalue.get()
    b = smevalue.get()
    c = smeeidvalue.get()
    d = supvalue.get()
    e = supeidvalue.get()
    f = slvalue.get()
    g = sleidvalue.get()
    h = olvalue.get()
    i = oleidvalue.get()

    file = openpyxl.load_workbook("Master1.xlsx")
    sheet = file.active
    sheet.cell(column = 1, row = sheet.max_row+1, value = a)
    sheet.cell(column = 2, row = sheet.max_row, value = b)
    sheet.cell(column = 3, row = sheet.max_row, value = c)
    sheet.cell(column = 4, row = sheet.max_row, value = d)
    sheet.cell(column = 5, row = sheet.max_row, value = e)
    sheet.cell(column = 6, row = sheet.max_row, value = f)
    sheet.cell(column = 7, row = sheet.max_row, value = g)
    sheet.cell(column = 8, row = sheet.max_row, value = h)
    sheet.cell(column = 9, row = sheet.max_row, value = i)
    file.save("Master1.xlsx")

#Defining frame 1
f1 = Frame(root, bg = "gray", borderwidth = 5, relief = SUNKEN)
f1.pack(side = LEFT, fill = Y)

#Defining labels & buttons in frame 1

acclogo = PhotoImage(file = "Accenture_logo_small.png")
Label(f1, image = acclogo).pack(side = TOP, pady = 2)

# selections = Label(f1, text = "Select from following options")
# selections.pack(pady = 5)

b1 = Button(f1, bg = "#0A7392", fg = "white", text = "Roster", font = ("Monotype Corsiva", 14),
            borderwidth = 5, relief = SUNKEN, padx = 31, command = rosterfunc)
b1.pack(pady = 35)

b2 = Button(f1, bg = "#0A7392", fg = "white", text = "Master Tracker", font = ("Monotype Corsiva", 14),
            borderwidth = 5, relief = SUNKEN, command = masterfunc)
b2.pack(pady = 35)

b3 = Button(f1, bg = "#0A7392", fg = "white", text = "Feedback", font = ("Monotype Corsiva", 14),
            borderwidth = 5, relief = SUNKEN, padx = 22, command = feedbackfunc)
b3.pack(pady = 35)

b4 = Button(f1, bg = "#0A7392", fg = "white", text = "About", font = ("Monotype Corsiva", 14),
            borderwidth = 5, relief = SUNKEN, padx = 32, command = aboutfunc)
b4.pack(pady = 35)



#Defining Frame 2
f2 = Frame(root, borderwidth = 15, relief = SUNKEN, bg = "green")
f2.pack(side = TOP, pady = 10)

#Defining labels in Frame 2
wlcmmsg = Label(f2, text = "Welcome to WFM Assistance Tool", fg = "black",
                font = ("Times New Roman", 24, "italic"))
wlcmmsg.pack(fill = X)



#Defining Frame 3
f3 = Frame(root, borderwidth = 5, relief = SUNKEN, pady = 25)
f3.pack(pady = 10, fill = X)

#Defining labels in Frame 3
empid = Label(f3, text = "Input employee ID", font = ("Times New Roman", 12), pady = 5, padx = 25)
entid = Label(f3, text = "Input enterprise ID", font = ("Times New Roman", 12), pady = 5, padx = 25)
empid.grid(row = 0, column = 1)
entid.grid(row = 1, column = 1)

#Defining data type for label values in Frame 3 and assigning them to entry box
empidvalue = IntVar()
entidvalue = StringVar()
empidentry = Entry(f3, textvariable = empidvalue)
entidentry = Entry(f3, textvariable = entidvalue)
empidentry.grid(row = 0, column = 2)
entidentry.grid(row = 1, column = 2)

#Defining Frame 4
f4 = Frame(root, borderwidth = 5, relief = SUNKEN)
f4.pack(anchor = "n", side = LEFT, fill = X, pady = 5, padx = 2)

#Defining labels & buttons inside Frame 4
Label(f4, text = "Current alignment", fg = "blue", font = ("Times New Roman", 14, "bold"), padx = 70, pady = 10).grid(row = 0, column = 3)
sme = Label(f4, text = "SME", font = ("Times New Roman", 10))
smeeid = Label(f4, text = "SME EID", font = ("Times New Roman", 10))
sup = Label(f4, text = "Supervisor", font = ("Times New Roman", 10))
supeid = Label(f4, text = "Supervisor EID", font = ("Times New Roman", 10))
sl = Label(f4, text = "Shift Lead", font = ("Times New Roman", 10))
sleid = Label(f4, text = "Shift Lead EID", font = ("Times New Roman", 10))
ol = Label(f4, text = "Ops Lead", font = ("Times New Roman", 10))
oleid = Label(f4, text = "Ops Lead EID", font = ("Times New Roman", 10))
Button(f4, text = "Retrieve data", fg = "white", bg = "#0A7392", padx = 5, pady = 5, command = retdatafunc).grid(row = 9, column = 2, padx = 10, pady = 25)
sme.grid(row = 1, column = 2)
smeeid.grid(row = 2, column = 2)
sup.grid(row = 3, column = 2)
supeid.grid(row = 4, column = 2)
sl.grid(row = 5, column = 2)
sleid.grid(row = 6, column = 2)
ol.grid(row = 7, column = 2)
oleid.grid(row = 8, column = 2)

#Defining data type for label values in Frame 4 and associating them to entry box
smevalue = StringVar()
smeeidvalue = IntVar()
supvalue = StringVar()
supeidvalue = IntVar()
slvalue = StringVar()
sleidvalue = IntVar()
olvalue = StringVar()
oleidvalue = IntVar()

smeentry = Entry(f4, textvariable = smevalue)
smeeidentry = Entry(f4, textvariable = smeeidvalue)
supentry = Entry(f4, textvariable = supvalue)
supeidentry = Entry(f4, textvariable = supeidvalue)
slentry = Entry(f4, textvariable = slvalue)
sleidentry = Entry(f4, textvariable = sleidvalue)
olentry = Entry(f4, textvariable = olvalue)
oleidentry = Entry(f4, textvariable = oleidvalue)

smeentry.grid(row = 1, column = 3)
smeeidentry.grid(row = 2, column = 3)
supentry.grid(row = 3, column = 3)
supeidentry.grid(row = 4, column = 3)
slentry.grid(row = 5, column = 3)
sleidentry.grid(row = 6, column = 3)
olentry.grid(row = 7, column = 3)
oleidentry.grid(row = 8, column = 3)

#Defining Frame 5
f5 = Frame(root, borderwidth = 5, relief = SUNKEN, padx = 10)
f5.pack(anchor = "n", side = RIGHT, fill = X, pady = 5, padx = 2)

#Defining labels & buttons inside Frame 5
Label(f5, text = "Request alignment change", fg = "blue", font = ("Times New Roman", 14, "bold"), padx = 30, pady = 10).grid(row = 0, column = 3)
sme = Label(f5, text = "SME", font = ("Times New Roman", 10))
smeeid = Label(f5, text = "SME EID", font = ("Times New Roman", 10))
sup = Label(f5, text = "Supervisor", font = ("Times New Roman", 10))
supeid = Label(f5, text = "Supervisor EID", font = ("Times New Roman", 10))
sl = Label(f5, text = "Shift Lead", font = ("Times New Roman", 10))
sleid = Label(f5, text = "Shift Lead EID", font = ("Times New Roman", 10))
ol = Label(f5, text = "Ops Lead", font = ("Times New Roman", 10))
oleid = Label(f5, text = "Ops Lead EID", font = ("Times New Roman", 10))
Button(f5, text = "Update data", fg = "white", bg = "#0A7392", padx = 5, pady = 5, command = updatafunc).grid(row = 9, column = 2, pady = 25)
sme.grid(row = 1, column = 2)
smeeid.grid(row = 2, column = 2)
sup.grid(row = 3, column = 2)
supeid.grid(row = 4, column = 2)
sl.grid(row = 5, column = 2)
sleid.grid(row = 6, column = 2)
ol.grid(row = 7, column = 2)
oleid.grid(row = 8, column = 2)

#Defining data type for label values in Frame 5 and associating them to entry box
smevalue = StringVar()
smeeidvalue = IntVar()
supvalue = StringVar()
supeidvalue = IntVar()
slvalue = StringVar()
sleidvalue = IntVar()
olvalue = StringVar()
oleidvalue = IntVar()

smeentry = Entry(f5, textvariable = smevalue)
smeeidentry = Entry(f5, textvariable = smeeidvalue)
supentry = Entry(f5, textvariable = supvalue)
supeidentry = Entry(f5, textvariable = supeidvalue)
slentry = Entry(f5, textvariable = slvalue)
sleidentry = Entry(f5, textvariable = sleidvalue)
olentry = Entry(f5, textvariable = olvalue)
oleidentry = Entry(f5, textvariable = oleidvalue)

smeentry.grid(row = 1, column = 3)
smeeidentry.grid(row = 2, column = 3)
supentry.grid(row = 3, column = 3)
supeidentry.grid(row = 4, column = 3)
slentry.grid(row = 5, column = 3)
sleidentry.grid(row = 6, column = 3)
olentry.grid(row = 7, column = 3)
oleidentry.grid(row = 8, column = 3)

# data1 = [{"Sap ID" : empidvalue.get(), "SME" : smevalue.get(), "SME EID" : smeeidvalue.get(),
#          "Supervisor" : supvalue.get(), "Sup Enterprise ID" : supeidvalue.get(), "Shift Lead" : slvalue.get(),
#          "Shift Lead EID" : sleidvalue.get(),"Ops Lead" : olvalue.get(), "Ops Lead EID" : oleidvalue.get()}]
# df1 = pd.DataFrame(data1)

#Creating an event loop
root.mainloop()